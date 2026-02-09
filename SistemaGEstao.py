import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from ofxparse import OfxParser
import pandas as pd
import json
from openpyxl import load_workbook

# Classes de modelo
class Cliente:
    def __init__(self, id, nome, endereco, valor):
        self.id = id
        self.nome = nome
        self.endereco = endereco
        self.valor = valor

class Recebimento:
    def __init__(self, cliente):
        self.cliente = cliente
        self.pagamentos = {mes: "Não pago" for mes in range(1, 13)}

    def marcar_pagamento(self, mes):
        self.pagamentos[mes] = "Pago" if self.pagamentos[mes] == "Não pago" else "Não pago"

# Funções auxiliares
def importar_ofx():
    arquivo = filedialog.askopenfilename(filetypes=[("OFX files", "*.ofx")])
    if not arquivo:
        return
    with open(arquivo) as f:
        ofx = OfxParser.parse(f)
    
    for transacao in ofx.account.statement.transactions:
        nome_cliente = transacao.payee
        valor = transacao.amount
        data = transacao.date
        
        cliente_existente = next((c for c in clientes if c.nome == nome_cliente), None)
        
        if cliente_existente:
            recebimento = next((r for r in recebimentos if r.cliente == cliente_existente), None)
            mes = data.month
            recebimento.marcar_pagamento(mes)
        else:
            cliente = Cliente(id=len(clientes)+1, nome=nome_cliente, endereco="", valor=valor)
            recebimento = Recebimento(cliente)
            mes = data.month
            recebimento.marcar_pagamento(mes)
            clientes.append(cliente)
            recebimentos.append(recebimento)
            pagos = sum(1 for v in recebimento.pagamentos.values() if v == "Pago")
            valor_total = pagos * cliente.valor
            tabela.insert("", "end", values=(cliente.nome, f"R$ {cliente.valor:,.2f}", *recebimento.pagamentos.values(), f"R$ {valor_total:,.2f}"))

def exportar_excel():
    nome_arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not nome_arquivo:
        return
    dados = []
    for recebimento in recebimentos:
        linha = [recebimento.cliente.nome, f"R$ {recebimento.cliente.valor:,.2f}"]
        linha.extend(recebimento.pagamentos.values())
        pagos = sum(1 for v in recebimento.pagamentos.values() if v == "Pago")
        valor_total = pagos * recebimento.cliente.valor
        linha.append(f"R$ {valor_total:,.2f}")
        dados.append(linha)
    
    df = pd.DataFrame(dados, columns=colunas)
    df.to_excel(nome_arquivo, engine='openpyxl', index=False)
    messagebox.showinfo("Sucesso", "Dados exportados com sucesso!")

def cadastrar_cliente():
    nome = simpledialog.askstring("Cadastro", "Nome do Cliente:")
    if nome:
        endereco = simpledialog.askstring("Cadastro", "Endereço do Cliente:")
        valor = simpledialog.askfloat("Cadastro", "Valor pago pelo Cliente:")
        if valor is not None:
            cliente = Cliente(id=len(clientes)+1, nome=nome, endereco=endereco, valor=valor)
            recebimento = Recebimento(cliente)
            clientes.append(cliente)
            recebimentos.append(recebimento)
            tabela.insert("", "end", values=(nome, f"R$ {valor:,.2f}", *recebimento.pagamentos.values(), f"R$ {0:,.2f}"))

def excluir_cliente():
    row_id = tabela.selection()[0]
    index = tabela.index(row_id)
    tabela.delete(row_id)
    del clientes[index]
    del recebimentos[index]

def excluir_clientes_duplicados():
    nomes_vistos = set()
    clientes_unicos = []
    recebimentos_unicos = []
    for cliente, recebimento in zip(clientes, recebimentos):
        if cliente.nome not in nomes_vistos:
            nomes_vistos.add(cliente.nome)
            clientes_unicos.append(cliente)
            recebimentos_unicos.append(recebimento)
    clientes.clear()
    clientes.extend(clientes_unicos)
    recebimentos.clear()
    recebimentos.extend(recebimentos_unicos)
    atualizar_tabela()

def importar_clientes_excel():
    arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not arquivo:
        return
    workbook = load_workbook(arquivo)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) != 3:
            continue
        nome, endereco, valor = row
        if valor is None:
            valor = 0.0
        cliente = Cliente(id=len(clientes)+1, nome=nome, endereco=endereco, valor=valor)
        recebimento = Recebimento(cliente)
        clientes.append(cliente)
        recebimentos.append(recebimento)
        pagos = sum(1 for v in recebimento.pagamentos.values() if v == "Pago")
        valor_total = pagos * cliente.valor
        tabela.insert("", "end", values=(cliente.nome, f"R$ {cliente.valor:,.2f}", *recebimento.pagamentos.values(), f"R$ {valor_total:,.2f}"))

def marcar_como_pago(event):
    row_id = tabela.selection()[0]
    col_id = tabela.identify_column(event.x)
    if col_id:
        mes = int(col_id[1:]) - 2
        if 1 <= mes <= 12:
            index = tabela.index(row_id)
            recebimento = recebimentos[index]
            recebimento.marcar_pagamento(mes)
            pagos = sum(1 for v in recebimento.pagamentos.values() if v == "Pago")
            valor_total = pagos * recebimento.cliente.valor
            tabela.item(row_id, values=(recebimento.cliente.nome, f"R$ {recebimento.cliente.valor:,.2f}", *recebimento.pagamentos.values(), f"R$ {valor_total:,.2f}"))

def salvar_dados():
    with open('dados.json', 'w') as f:
        data = {
            'clientes': [{'id': c.id, 'nome': c.nome, 'endereco': c.endereco, 'valor': c.valor} for c in clientes],
            'recebimentos': [{'cliente_id': r.cliente.id, 'pagamentos': r.pagamentos} for r in recebimentos]
        }
        json.dump(data, f)

def carregar_dados():
    try:
        with open('dados.json', 'r') as f:
            data = json.load(f)
            for cliente_data in data['clientes']:
                valor = cliente_data['valor'] if cliente_data['valor'] is not None else 0.0
                cliente = Cliente(cliente_data['id'], cliente_data['nome'], cliente_data['endereco'], valor)
                clientes.append(cliente)
            for recebimento_data in data['recebimentos']:
                cliente = next(c for c in clientes if c.id == recebimento_data['cliente_id'])
                recebimento = Recebimento(cliente)
                recebimento.pagamentos = recebimento_data['pagamentos']
                recebimentos.append(recebimento)
                pagos = sum(1 for v in recebimento.pagamentos.values() if v == "Pago")
                valor_total = pagos * cliente.valor
                tabela.insert("", "end", values=(cliente.nome, f"R$ {cliente.valor:,.2f}", *recebimento.pagamentos.values(), f"R$ {valor_total:,.2f}"))
    except FileNotFoundError:
        pass

def buscar_cliente():
    nome = simpledialog.askstring("Buscar Cliente", "Nome do Cliente:")
    if nome:
        for i, cliente in enumerate(clientes):
            if cliente.nome.lower() == nome.lower():
                tabela.selection_set(tabela.get_children()[i])
                tabela.see(tabela.get_children()[i])
                break
        else:
            messagebox.showinfo("Info", "Cliente não encontrado.")

def atualizar_cliente():
    row_id = tabela.selection()[0]
    index = tabela.index(row_id)
    cliente = clientes[index]
    nome = simpledialog.askstring("Atualizar Cliente", "Nome do Cliente:", initialvalue=cliente.nome)
    endereco = simpledialog.askstring("Atualizar Cliente", "Endereço do Cliente:", initialvalue=cliente.endereco)
    valor = simpledialog.askfloat("Atualizar Cliente", "Valor pago pelo Cliente:", initialvalue=cliente.valor)
    if nome and endereco and valor is not None:
        cliente.nome = nome
        cliente.endereco = endereco
        cliente.valor = valor
        recebimento = recebimentos[index]
        pagos = sum(1 for v in recebimento.pagamentos.values() if v == "Pago")
        valor_total = pagos * cliente.valor
        tabela.item(row_id, values=(nome, f"R$ {valor:,.2f}", *recebimento.pagamentos.values(), f"R$ {valor_total:,.2f}"))

def atualizar_tabela():
    for row in tabela.get_children():
        tabela.delete(row)
    for recebimento in recebimentos:
        pagos = sum(1 for v in recebimento.pagamentos.values() if v == "Pago")
        valor_total = pagos * (recebimento.cliente.valor if recebimento.cliente.valor is not None else 0.0)
        tabela.insert("", "end", values=(recebimento.cliente.nome, f"R$ {recebimento.cliente.valor:,.2f}", *recebimento.pagamentos.values(), f"R$ {valor_total:,.2f}"))

# Inicialização dos dados e da interface
clientes = []
recebimentos = []

root = tk.Tk()
root.title("Sistema de Gestão de Clientes")

frame = tk.Frame(root)
frame.pack(fill=tk.BOTH, expand=True)

colunas = ["Cliente", "Valor", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez", "Total"]
tabela = ttk.Treeview(frame, columns=colunas, show="headings")
tabela.pack(fill=tk.BOTH, expand=True)

for col in colunas:
    tabela.heading(col, text=col)
    tabela.column(col, minwidth=0, width=80, stretch=tk.NO)

tabela.bind("<Double-1>", marcar_como_pago)

botoes_frame = tk.Frame(root)
botoes_frame.pack(fill=tk.X)

tk.Button(botoes_frame, text="Importar OFX", command=importar_ofx).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Exportar Excel", command=exportar_excel).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Cadastrar Cliente", command=cadastrar_cliente).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Excluir Cliente", command=excluir_cliente).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Importar Clientes Excel", command=importar_clientes_excel).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Excluir Clientes Duplicados", command=excluir_clientes_duplicados).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Buscar Cliente", command=buscar_cliente).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Atualizar Cliente", command=atualizar_cliente).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Salvar Dados", command=salvar_dados).pack(side=tk.LEFT)
tk.Button(botoes_frame, text="Carregar Dados", command=carregar_dados).pack(side=tk.LEFT)

carregar_dados()
root.mainloop()
