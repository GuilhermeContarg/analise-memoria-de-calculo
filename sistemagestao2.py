import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from openpyxl import load_workbook
import pandas as pd
import json

# Classes de modelo
class Cliente:
    def __init__(self, nome, cnpj=None, contato=None, valor_mensalidade=None):
        self.nome = nome
        self.cnpj = cnpj
        self.contato = contato
        self.valor_mensalidade = valor_mensalidade if valor_mensalidade is not None else 0

    def to_dict(self):
        return {
            "Nome": self.nome,
            "CNPJ": self.cnpj,
            "Contato": self.contato,
            "Valor Mensalidade": self.valor_mensalidade
        }

class Recebimento:
    def __init__(self, cliente):
        self.cliente = cliente
        self.pagamentos = {mes: "Não pago" for mes in range(1, 13)}

    def marcar_pagamento(self, mes):
        self.pagamentos[mes] = "Pago"

# Funções relacionadas à interface de usuário
def cadastrar_cliente():
    nome = simpledialog.askstring("Cadastro de Cliente", "Nome do Cliente:")
    if nome:
        cliente = Cliente(nome)
        clientes.append(cliente)
        atualizar_tabela_clientes()

def importar_clientes_excel():
    arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if arquivo:
        workbook = load_workbook(arquivo)
        sheet = workbook.active

        colunas = {
            "Nome": "A",
            "CNPJ": "B",
            "Contato": "C",
            "Valor Mensalidade": "D"
        }

        dados_clientes = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[colunas["Nome"]]
            cnpj = row[colunas["CNPJ"]]
            contato = row[colunas["Contato"]]
            valor_mensalidade = row[colunas["Valor Mensalidade"]]
            cliente = Cliente(nome, cnpj, contato, valor_mensalidade)
            clientes.append(cliente)

        atualizar_tabela_clientes()

def exportar_excel():
    dados = [cliente.to_dict() for cliente in clientes]
    df = pd.DataFrame(dados)
    nome_arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if nome_arquivo:
        df.to_excel(nome_arquivo, index=False)
        messagebox.showinfo("Sucesso", "Dados exportados com sucesso!")

def carregar_dados():
    try:
        with open('dados.json', 'r') as f:
            data = json.load(f)
            for cliente_data in data['clientes']:
                cliente = Cliente(**cliente_data)
                clientes.append(cliente)
    except FileNotFoundError:
        pass

def salvar_dados():
    with open('dados.json', 'w') as f:
        data = {'clientes': [cliente.to_dict() for cliente in clientes]}
        json.dump(data, f)

def atualizar_tabela_clientes():
    for i in tabela_clientes.get_children():
        tabela_clientes.delete(i)
    for cliente in clientes:
        tabela_clientes.insert("", "end", values=(cliente.nome, cliente.cnpj, cliente.contato, f"R$ {cliente.valor_mensalidade:,.2f}"))

# Criação da GUI
root = tk.Tk()
root.title("Sistema de Gestão")

frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
frame.columnconfigure(0, weight=1)
frame.rowconfigure(3, weight=1)

ttk.Button(frame, text="Cadastrar Cliente", command=cadastrar_cliente).grid(row=0, column=0, sticky=tk.W, pady=5)
ttk.Button(frame, text="Importar Clientes (Excel)", command=importar_clientes_excel).grid(row=1, column=0, sticky=tk.W, pady=5)
ttk.Button(frame, text="Exportar Excel", command=exportar_excel).grid(row=2, column=0, sticky=tk.W, pady=5)

container = ttk.Frame(frame)
container.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
container.columnconfigure(0, weight=1)
container.rowconfigure(0, weight=1)

colunas_tabela_clientes = ["Nome", "CNPJ", "Contato", "Valor Mensalidade"]
tabela_clientes = ttk.Treeview(container, columns=colunas_tabela_clientes, show="headings")
for col in colunas_tabela_clientes:
    tabela_clientes.heading(col, text=col)
tabela_clientes.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

scrollbar_vertical = ttk.Scrollbar(container, orient="vertical", command=tabela_clientes.yview)
scrollbar_vertical.grid(row=0, column=1, sticky=(tk.N, tk.S))
tabela_clientes.configure(yscrollcommand=scrollbar_vertical.set)

scrollbar_horizontal = ttk.Scrollbar(container, orient="horizontal", command=tabela_clientes.xview)
scrollbar_horizontal.grid(row=1, column=0, sticky=(tk.W, tk.E))
tabela_clientes.configure(xscrollcommand=scrollbar_horizontal.set)

clientes = []
carregar_dados()

root.mainloop()
salvar_dados()
